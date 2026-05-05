"""
Testes de integração — fluxo completo de batch.

Estratégia de mock:
  - Parsers DXF e PDF rodam de verdade (ezdxf / PyMuPDF).
  - Apenas NLPProcessor.process() é mockado para evitar chamadas à API.

Cada classe testa um nível diferente da pilha:

  TestParsePhase         parse_single com parsers reais → verifica conteúdo do raw_summary
  TestLLMPhase           run_llm_batch com NLPProcessor mockado → verifica mutações de FileJob
  TestDataContinuity     garante que o texto real do arquivo chega ao LLM
  TestCombinedPipeline   bytes → BatchResult → combined_tables, vários tipos de arquivo
  TestExportIntegration  BatchResult → Excel verificável + JSON serializável
  TestHistoryIntegration batch + HistoryEntry → identidade de objetos e deduplicação
  TestEdgeCases          batch vazio, arquivo inválido, mistura de status
"""

from __future__ import annotations

import json
import math
import os
import tempfile
from pathlib import Path
from typing import Any
from unittest.mock import patch

import openpyxl
import pandas as pd
import pytest

from src.batch.batch_processor import BatchProcessor, BatchResult, FileJob
from src.history import HistoryEntry, append_entry
from src.utils.exporters import export_batch_to_excel, result_to_dataframe

pytestmark = pytest.mark.integration


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_processor(**kwargs) -> BatchProcessor:
    return BatchProcessor(**kwargs)


def _patch_llm(return_value: dict):
    """Context manager: mocka NLPProcessor.process() com resposta fixa."""
    return patch(
        "src.batch.batch_processor.NLPProcessor",
        **{"return_value.process.return_value": return_value},
    )


# ---------------------------------------------------------------------------
# 1. Fase de parse — parsers reais, sem mock
# ---------------------------------------------------------------------------

class TestParsePhase:
    """Verifica que parse_single produz FileJob com conteúdo real e correto."""

    def test_dxf_job_status_ready(self, dxf_sala: bytes):
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h_sala")
        assert job.status == "ready"
        assert job.file_type == "dxf"
        assert job.raw_summary is not None

    def test_dxf_raw_summary_contains_text_content(self, dxf_sala: bytes):
        """O texto 'SALA DE ESTAR' inserido no DXF deve aparecer no raw_summary."""
        job = _make_processor().parse_single("sala.dxf", dxf_sala, "h")
        assert "SALA DE ESTAR" in job.raw_summary

    def test_dxf_raw_summary_contains_area_text(self, dxf_sala: bytes):
        """Texto de área inserido no DXF deve ser preservado no summary."""
        job = _make_processor().parse_single("sala.dxf", dxf_sala, "h")
        assert "AREA=20.00M2" in job.raw_summary

    def test_dxf_raw_summary_contains_closed_polyline_section(self, dxf_sala: bytes):
        """A seção de polilínhas fechadas deve estar presente."""
        job = _make_processor().parse_single("sala.dxf", dxf_sala, "h")
        assert "POLILÍNHAS FECHADAS" in job.raw_summary

    def test_dxf_raw_summary_contains_block_section(self, dxf_sala: bytes):
        """Blocos (porta) devem aparecer na seção de inserções."""
        job = _make_processor().parse_single("sala.dxf", dxf_sala, "h")
        assert "BLOCOS" in job.raw_summary
        assert "P-90-SALA" in job.raw_summary

    def test_dxf_layer_filter_restricts_content(self, dxf_sala: bytes):
        """Com filtro de layer, apenas entidades nessa layer devem ser extraídas."""
        job_all   = _make_processor().parse_single("sala.dxf", dxf_sala, "h_all")
        job_room  = _make_processor(dxf_layers=["A-ROOM"]).parse_single(
            "sala.dxf", dxf_sala, "h_room"
        )
        # Sem filtro tem mais conteúdo que com filtro
        assert len(job_all.raw_summary) > len(job_room.raw_summary)
        # Com filtro de A-ROOM, textos de A-TEXT não aparecem
        assert "SALA DE ESTAR" not in job_room.raw_summary

    def test_dxf_units_appear_in_summary(self, dxf_sala: bytes):
        """Unidades do documento (Metros) devem constar no summary."""
        job = _make_processor().parse_single("sala.dxf", dxf_sala, "h")
        assert "Metros" in job.raw_summary

    def test_pdf_job_status_ready(self, pdf_memorial: bytes):
        bp = _make_processor()
        job = bp.parse_single("memorial.pdf", pdf_memorial, "h_pdf")
        assert job.status == "ready"
        assert job.file_type == "pdf"
        assert job.raw_summary is not None

    def test_pdf_raw_summary_contains_page_section(self, pdf_memorial: bytes):
        job = _make_processor().parse_single("memorial.pdf", pdf_memorial, "h")
        assert "PÁGINA 1" in job.raw_summary
        assert "PÁGINA 2" in job.raw_summary

    def test_pdf_raw_summary_contains_extracted_text(self, pdf_memorial: bytes):
        """Texto inserido no PDF deve aparecer (ao menos parcialmente) no summary."""
        job = _make_processor().parse_single("memorial.pdf", pdf_memorial, "h")
        summary_lower = job.raw_summary.lower()
        # Ao menos um dos termos chave do PDF deve estar presente
        assert any(kw in summary_lower for kw in ("sala", "dormitório", "memorial", "20.50"))

    def test_parse_records_duration(self, dxf_sala: bytes):
        job = _make_processor().parse_single("sala.dxf", dxf_sala, "h")
        assert job.parse_duration_s > 0.0

    def test_corrupted_dxf_produces_error_job(self):
        job = _make_processor().parse_single("bad.dxf", b"%%%NOT A DXF%%%", "h_bad")
        assert job.status == "error"
        assert job.error is not None
        assert job.raw_summary is None
        assert job.parse_result is None

    def test_unsupported_extension_produces_error_job(self, dxf_sala: bytes):
        job = _make_processor().parse_single("arquivo.dwg", dxf_sala, "h_dwg")
        assert job.status == "error"
        assert "DWG" in job.error or "suportado" in job.error.lower()


# ---------------------------------------------------------------------------
# 2. Fase LLM — mutações de FileJob, callbacks, isolamento de erros
# ---------------------------------------------------------------------------

class TestLLMPhase:
    """Verifica o comportamento de run_llm_batch com NLPProcessor mockado."""

    def test_successful_job_becomes_done(self, dxf_sala, llm_resp_ambientes):
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h")

        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch([job], "extraia ambientes")

        assert job.status == "done"
        assert job.llm_result is not None
        assert result.successful == [job]

    def test_llm_result_stored_on_job(self, dxf_sala, llm_resp_ambientes):
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h")

        with _patch_llm(llm_resp_ambientes):
            bp.run_llm_batch([job], "q")

        assert "ambientes" in job.llm_result
        assert job.llm_result["confianca"] == "alta"

    def test_llm_error_sets_error_status(self, dxf_sala):
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h")

        with patch("src.batch.batch_processor.NLPProcessor") as mock_cls:
            mock_cls.return_value.process.side_effect = RuntimeError("timeout")
            result = bp.run_llm_batch([job], "q")

        assert job.status == "error"
        assert "LLM" in job.error
        assert result.failed == [job]

    def test_one_llm_error_does_not_block_remaining(self, dxf_sala, dxf_quarto, llm_resp_ambientes):
        bp = _make_processor()
        job1 = bp.parse_single("sala.dxf",   dxf_sala,   "h1")
        job2 = bp.parse_single("quarto.dxf", dxf_quarto, "h2")

        call_n = 0
        def side_effect(raw_text, user_query):
            nonlocal call_n
            call_n += 1
            if call_n == 1:
                raise RuntimeError("API rate limit")
            return llm_resp_ambientes

        with patch("src.batch.batch_processor.NLPProcessor") as mock_cls:
            mock_cls.return_value.process.side_effect = side_effect
            result = bp.run_llm_batch([job1, job2], "q")

        assert job1.status == "error"
        assert job2.status == "done"
        assert len(result.successful) == 1
        assert len(result.failed) == 1

    def test_pending_job_is_skipped(self, dxf_sala, llm_resp_ambientes):
        bp = _make_processor()
        ready_job   = bp.parse_single("sala.dxf", dxf_sala, "h1")
        pending_job = FileJob(filename="pending.dxf", file_hash="h2",
                              size_bytes=0, status="pending")

        with patch("src.batch.batch_processor.NLPProcessor") as mock_cls:
            mock_cls.return_value.process.return_value = llm_resp_ambientes
            bp.run_llm_batch([ready_job, pending_job], "q")
            assert mock_cls.return_value.process.call_count == 1

        assert pending_job.status == "pending"

    def test_progress_callback_order(self, dxf_sala, dxf_quarto, llm_resp_ambientes):
        bp = _make_processor()
        job1 = bp.parse_single("a.dxf", dxf_sala,   "h1")
        job2 = bp.parse_single("b.dxf", dxf_quarto, "h2")

        calls: list[tuple] = []
        with _patch_llm(llm_resp_ambientes):
            bp.run_llm_batch([job1, job2], "q",
                             on_progress=lambda c, t, f: calls.append((c, t, f)))

        assert calls[0]  == (0, 2, "a.dxf")   # antes do 1º arquivo
        assert calls[1]  == (1, 2, "b.dxf")   # antes do 2º arquivo
        assert calls[-1] == (2, 2, "")         # sinal de conclusão

    def test_llm_duration_recorded(self, dxf_sala, llm_resp_ambientes):
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h")
        with _patch_llm(llm_resp_ambientes):
            bp.run_llm_batch([job], "q")
        assert job.llm_duration_s >= 0.0


# ---------------------------------------------------------------------------
# 3. Continuidade de dados — o texto real chega ao LLM
# ---------------------------------------------------------------------------

class TestDataContinuity:
    """
    Verifica que o raw_summary produzido pelo parser é exatamente o texto
    que chega como argumento para NLPProcessor.process().
    """

    def test_raw_summary_is_passed_verbatim_to_llm(self, dxf_sala):
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h")

        received: list[dict] = []
        def capture(raw_text, user_query):
            received.append({"raw_text": raw_text, "query": user_query})
            return {}

        with patch("src.batch.batch_processor.NLPProcessor") as mock_cls:
            mock_cls.return_value.process.side_effect = capture
            bp.run_llm_batch([job], "extraia ambientes e áreas")

        assert len(received) == 1
        assert received[0]["raw_text"] == job.raw_summary
        assert received[0]["query"] == "extraia ambientes e áreas"

    def test_summary_sent_to_llm_contains_real_dxf_text(self, dxf_sala):
        """O texto que chega ao LLM deve conter 'SALA DE ESTAR' — vindo do DXF real."""
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h")

        received_text: list[str] = []
        with patch("src.batch.batch_processor.NLPProcessor") as mock_cls:
            mock_cls.return_value.process.side_effect = (
                lambda raw_text, **kw: received_text.append(raw_text) or {}
            )
            bp.run_llm_batch([job], "q")

        assert received_text, "LLM não foi chamado"
        assert "SALA DE ESTAR" in received_text[0]

    def test_pdf_text_reaches_llm(self, pdf_memorial):
        bp = _make_processor()
        job = bp.parse_single("memorial.pdf", pdf_memorial, "h")

        received_text: list[str] = []
        with patch("src.batch.batch_processor.NLPProcessor") as mock_cls:
            mock_cls.return_value.process.side_effect = (
                lambda raw_text, **kw: received_text.append(raw_text) or {}
            )
            bp.run_llm_batch([job], "q")

        assert received_text
        summary_lower = received_text[0].lower()
        assert any(kw in summary_lower for kw in ("sala", "dormitório", "20.50", "memorial"))

    def test_multiple_files_each_get_own_summary(self, dxf_sala, dxf_quarto):
        """Dois arquivos distintos devem enviar summaries distintos ao LLM."""
        bp = _make_processor()
        job1 = bp.parse_single("sala.dxf",   dxf_sala,   "h1")
        job2 = bp.parse_single("quarto.dxf", dxf_quarto, "h2")

        received: list[str] = []
        with patch("src.batch.batch_processor.NLPProcessor") as mock_cls:
            mock_cls.return_value.process.side_effect = (
                lambda raw_text, **kw: received.append(raw_text) or {}
            )
            bp.run_llm_batch([job1, job2], "q")

        assert len(received) == 2
        assert received[0] != received[1], "Os dois arquivos geraram o mesmo summary"
        assert "SALA DE ESTAR"  in received[0]
        assert "DORMITÓRIO 01" in received[1]


# ---------------------------------------------------------------------------
# 4. Pipeline combinado — bytes → BatchResult → tabelas mescladas
# ---------------------------------------------------------------------------

class TestCombinedPipeline:
    """Testa o pipeline completo sem separar fases."""

    def test_two_dxf_combined_table_has_arquivo_column(
        self, dxf_sala, dxf_quarto, llm_resp_ambientes
    ):
        bp = _make_processor()
        jobs = [
            bp.parse_single("sala.dxf",   dxf_sala,   "h1"),
            bp.parse_single("quarto.dxf", dxf_quarto, "h2"),
        ]
        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch(jobs, "q")

        df = result.combined_tables()["ambientes"]
        assert "Arquivo" in df.columns
        assert set(df["Arquivo"]) == {"sala", "quarto"}

    def test_combined_table_row_count_equals_sum_of_all_files(
        self, dxf_sala, dxf_quarto, llm_resp_ambientes
    ):
        """2 arquivos × 2 ambientes/arquivo = 4 linhas na tabela combinada."""
        bp = _make_processor()
        jobs = [
            bp.parse_single("sala.dxf",   dxf_sala,   "h1"),
            bp.parse_single("quarto.dxf", dxf_quarto, "h2"),
        ]
        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch(jobs, "q")

        df = result.combined_tables()["ambientes"]
        assert len(df) == 4

    def test_dxf_and_pdf_in_same_batch(
        self, dxf_sala, pdf_memorial, llm_resp_ambientes
    ):
        """DXF e PDF devem ser processados no mesmo lote sem erros."""
        bp = _make_processor()
        jobs = [
            bp.parse_single("sala.dxf",    dxf_sala,    "h1"),
            bp.parse_single("memorial.pdf", pdf_memorial, "h2"),
        ]
        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch(jobs, "q")

        assert len(result.successful) == 2
        assert len(result.failed) == 0

    def test_mixed_valid_and_invalid_files(self, dxf_sala, llm_resp_ambientes):
        """Arquivo inválido não deve impedir os demais."""
        bp = _make_processor()
        good = bp.parse_single("sala.dxf", dxf_sala,         "h_good")
        bad  = bp.parse_single("lixo.dxf", b"ARQUIVO LIXO",  "h_bad")

        assert good.status == "ready"
        assert bad.status  == "error"

        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch([good, bad], "q")

        assert result.successful == [good]
        assert result.failed     == [bad]

    def test_success_rate_with_partial_llm_failure(self, dxf_sala, dxf_quarto):
        bp = _make_processor()
        job1 = bp.parse_single("sala.dxf",   dxf_sala,   "h1")
        job2 = bp.parse_single("quarto.dxf", dxf_quarto, "h2")

        with patch("src.batch.batch_processor.NLPProcessor") as mock_cls:
            mock_cls.return_value.process.side_effect = [
                {"ambientes": []},
                RuntimeError("erro"),
            ]
            result = bp.run_llm_batch([job1, job2], "q")

        assert math.isclose(result.success_rate, 0.5)

    def test_batch_result_to_dict_is_json_serializable(
        self, dxf_sala, llm_resp_ambientes
    ):
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h")
        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch([job], "q")

        d = result.to_dict()
        # Não deve levantar exceção
        serialized = json.dumps(d, ensure_ascii=False)
        reloaded   = json.loads(serialized)

        assert reloaded["summary"]["successful"] == 1
        assert reloaded["results"][0]["filename"] == "sala.dxf"

    def test_multiambiente_dxf_summary_covers_all_rooms(self, dxf_multiambiente):
        """DXF com 3 ambientes deve gerar summary com todos os textos."""
        job = _make_processor().parse_single("multi.dxf", dxf_multiambiente, "h")
        assert job.status == "ready"
        for room in ("SALA DE ESTAR", "DORMITÓRIO 01", "COZINHA"):
            assert room in job.raw_summary


# ---------------------------------------------------------------------------
# 5. Exportação — Excel verificável, JSON consistente
# ---------------------------------------------------------------------------

class TestExportIntegration:
    """Verifica que os arquivos gerados são válidos e têm a estrutura esperada."""

    @pytest.fixture
    def batch_result_dois_arquivos(self, dxf_sala, dxf_quarto, llm_resp_ambientes):
        bp = _make_processor()
        jobs = [
            bp.parse_single("planta01.dxf", dxf_sala,   "h1"),
            bp.parse_single("planta02.dxf", dxf_quarto, "h2"),
        ]
        with _patch_llm(llm_resp_ambientes):
            return bp.run_llm_batch(jobs, "extraia ambientes")

    def test_excel_file_is_created(self, batch_result_dois_arquivos, tmp_path):
        br = batch_result_dois_arquivos
        out = tmp_path / "lote.xlsx"
        export_batch_to_excel(
            br.combined_tables(),
            [(j.filename, j.llm_result) for j in br.successful if j.llm_result],
            out,
        )
        assert out.exists()
        assert out.stat().st_size > 0

    def test_excel_has_combined_sheet(self, batch_result_dois_arquivos, tmp_path):
        br = batch_result_dois_arquivos
        out = tmp_path / "lote.xlsx"
        export_batch_to_excel(
            br.combined_tables(),
            [(j.filename, j.llm_result) for j in br.successful if j.llm_result],
            out,
        )
        wb = openpyxl.load_workbook(out)
        assert any("COMB" in name for name in wb.sheetnames)

    def test_excel_has_per_file_sheets(self, batch_result_dois_arquivos, tmp_path):
        br = batch_result_dois_arquivos
        out = tmp_path / "lote.xlsx"
        export_batch_to_excel(
            br.combined_tables(),
            [(j.filename, j.llm_result) for j in br.successful if j.llm_result],
            out,
        )
        wb = openpyxl.load_workbook(out)
        sheet_names = wb.sheetnames
        assert any("planta01" in name for name in sheet_names)
        assert any("planta02" in name for name in sheet_names)

    def test_excel_combined_sheet_has_arquivo_column(
        self, batch_result_dois_arquivos, tmp_path
    ):
        br = batch_result_dois_arquivos
        out = tmp_path / "lote.xlsx"
        export_batch_to_excel(
            br.combined_tables(),
            [(j.filename, j.llm_result) for j in br.successful if j.llm_result],
            out,
        )
        wb = openpyxl.load_workbook(out)
        comb_sheet = next(s for s in wb.sheetnames if "COMB" in s)
        headers = [cell.value for cell in wb[comb_sheet][1]]
        assert "Arquivo" in headers

    def test_excel_combined_sheet_row_count(
        self, batch_result_dois_arquivos, tmp_path
    ):
        br = batch_result_dois_arquivos
        out = tmp_path / "lote.xlsx"
        export_batch_to_excel(
            br.combined_tables(),
            [(j.filename, j.llm_result) for j in br.successful if j.llm_result],
            out,
        )
        wb = openpyxl.load_workbook(out)
        comb_sheet = next(s for s in wb.sheetnames if "COMB" in s)
        ws = wb[comb_sheet]
        # 1 cabeçalho + 4 linhas (2 ambientes × 2 arquivos)
        assert ws.max_row == 5

    def test_excel_sheet_names_are_valid(self, batch_result_dois_arquivos, tmp_path):
        """Nenhum nome de aba pode ter caracteres proibidos pelo Excel."""
        import re
        forbidden = re.compile(r"[/\\?*\[\]']")
        br = batch_result_dois_arquivos
        out = tmp_path / "lote.xlsx"
        export_batch_to_excel(
            br.combined_tables(),
            [(j.filename, j.llm_result) for j in br.successful if j.llm_result],
            out,
        )
        wb = openpyxl.load_workbook(out)
        for name in wb.sheetnames:
            assert not forbidden.search(name), f"Nome inválido: {name!r}"
            assert len(name) <= 31, f"Nome muito longo: {name!r}"


# ---------------------------------------------------------------------------
# 6. Integração com histórico
# ---------------------------------------------------------------------------

class TestHistoryIntegration:
    """
    Verifica que o BatchResult armazenado no histórico é o mesmo objeto
    (identidade) e que as mutações do pipeline são visíveis pelo histórico.
    """

    def test_history_entry_references_same_batch_result(
        self, dxf_sala, llm_resp_ambientes
    ):
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h")
        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch([job], "q")

        history: list[HistoryEntry] = []
        entry = HistoryEntry(
            query="q", files=["sala.dxf"],
            successful=len(result.successful), failed=len(result.failed),
            batch_result=result,
        )
        append_entry(history, entry)

        assert history[0].batch_result is result

    def test_history_entry_jobs_are_same_objects(
        self, dxf_sala, llm_resp_ambientes
    ):
        """O FileJob dentro do BatchResult guardado no histórico é o mesmo."""
        bp = _make_processor()
        job = bp.parse_single("sala.dxf", dxf_sala, "h")
        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch([job], "q")

        history: list[HistoryEntry] = []
        append_entry(history, HistoryEntry(
            query="q", files=["sala.dxf"],
            successful=1, failed=0, batch_result=result,
        ))

        assert history[0].batch_result.jobs[0] is job

    def test_restored_result_preserves_combined_tables(
        self, dxf_sala, dxf_quarto, llm_resp_ambientes
    ):
        """combined_tables() deve funcionar sobre um resultado restaurado do histórico."""
        bp = _make_processor()
        jobs = [
            bp.parse_single("a.dxf", dxf_sala,   "h1"),
            bp.parse_single("b.dxf", dxf_quarto, "h2"),
        ]
        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch(jobs, "q")

        history: list[HistoryEntry] = []
        append_entry(history, HistoryEntry(
            query="q", files=["a.dxf", "b.dxf"],
            successful=2, failed=0, batch_result=result,
        ))

        # Simula restauração: usa o batch_result do histórico
        restored = history[0].batch_result
        combined = restored.combined_tables()

        assert "ambientes" in combined
        assert len(combined["ambientes"]) == 4
        assert set(combined["ambientes"]["Arquivo"]) == {"a", "b"}

    def test_history_deduplication_preserves_latest_result(
        self, dxf_sala, llm_resp_ambientes
    ):
        """Dois runs com mesma query+arquivo devem resultar em apenas 1 entrada (a mais nova)."""
        bp = _make_processor()

        history: list[HistoryEntry] = []

        for _ in range(2):
            job = bp.parse_single("sala.dxf", dxf_sala, "h")
            with _patch_llm(llm_resp_ambientes):
                result = bp.run_llm_batch([job], "extraia ambientes")
            append_entry(history, HistoryEntry(
                query="extraia ambientes",
                files=["sala.dxf"],
                successful=1, failed=0,
                batch_result=result,
            ))

        assert len(history) == 1
        # A entrada deve ser a segunda (mais recente)
        assert history[0].batch_result is result


# ---------------------------------------------------------------------------
# 7. Edge cases
# ---------------------------------------------------------------------------

class TestEdgeCases:
    def test_empty_jobs_list(self):
        result = _make_processor().run_llm_batch([], "q")
        assert result.total == 0
        assert result.successful == []
        assert result.failed == []
        assert result.combined_tables() == {}

    def test_all_jobs_have_parse_errors(self):
        bp = _make_processor()
        bad1 = bp.parse_single("a.dxf", b"LIXO", "h1")
        bad2 = bp.parse_single("b.dxf", b"LIXO", "h2")

        with patch("src.batch.batch_processor.NLPProcessor") as mock_cls:
            result = bp.run_llm_batch([bad1, bad2], "q")
            mock_cls.return_value.process.assert_not_called()

        assert result.successful == []
        assert len(result.failed) == 2

    def test_to_dict_with_failed_jobs(self):
        bp = _make_processor()
        bad = bp.parse_single("bad.dxf", b"LIXO", "h")
        result = bp.run_llm_batch([bad], "q")

        d = result.to_dict()
        assert d["summary"]["failed"] == 1
        assert d["results"][0]["error"] is not None
        assert d["results"][0]["llm_result"] is None

    def test_combined_tables_empty_when_no_successful_jobs(self):
        bp = _make_processor()
        bad = bp.parse_single("bad.dxf", b"LIXO", "h")
        result = bp.run_llm_batch([bad], "q")

        assert result.combined_tables() == {}

    def test_large_batch_all_succeed(self, dxf_sala, llm_resp_ambientes):
        """10 arquivos DXF idênticos (hash diferente) devem todos completar."""
        bp = _make_processor()
        jobs = [
            bp.parse_single(f"planta{i:02d}.dxf", dxf_sala, f"hash_{i}")
            for i in range(10)
        ]
        assert all(j.status == "ready" for j in jobs)

        with _patch_llm(llm_resp_ambientes):
            result = bp.run_llm_batch(jobs, "q")

        assert len(result.successful) == 10
        assert result.success_rate == 1.0

        df = result.combined_tables()["ambientes"]
        assert len(df) == 20   # 2 ambientes × 10 arquivos
